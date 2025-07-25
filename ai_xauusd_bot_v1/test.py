import os
import re
import threading
import tkinter as tk
from tkinter import filedialog, BOTH, LEFT, RIGHT, X, Y, TOP, BOTTOM
from tkinterdnd2 import TkinterDnD, DND_FILES
from openpyxl import load_workbook
from ttkbootstrap import Style
from ttkbootstrap import ttk
from ttkbootstrap.dialogs import Messagebox
from ttkbootstrap.widgets import Meter, Progressbar, Label
from translator import Translator
from translation_dict import translate_dict

translator = Translator(translate_dict)

class ExcelTranslateApp:
    def __init__(self, root):
        self.root = root
        self.drop_area = None
        self.drop_text = None
        self.root.title("✨ Excel Translator - Dịch giữ định dạng ✨")
        self.root.geometry("750x350")
        self.root.resizable(False, False)

        self.style = Style("flatly")  # 💡 Thêm dòng này

        self.file_paths = []
        self.selected_sheet = tk.StringVar()

        self.setup_widgets()

    def setup_widgets(self):
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Tiêu đề
        ttk.Label(main_frame, text="📁 Dịch nhiều file Excel Nhật - Việt", font=("Segoe UI", 16, "bold")).pack(pady=10)

        # Vùng kéo thả
        self.drop_area = ttk.LabelFrame(main_frame, text="🧲 Kéo thả file Excel", bootstyle="info")
        self.drop_area.pack(fill=tk.BOTH, pady=10, ipadx=10, ipady=10, expand=True)
        self.drop_area.pack_propagate(False)

        self.drop_text = ttk.Label(self.drop_area, text="(Thả tối đa 10 file .xlsx vào đây)", font=("Segoe UI", 11))
        self.drop_text.pack(expand=True)

        # Kích hoạt kéo thả
        # self.drop_area.drop_target_register(DND_FILES)
        # self.drop_area.dnd_bind('<<Drop>>', self.on_drop_files)

        # Nút chọn file
        self.choose_button = ttk.Button(main_frame, text="📂 Chọn file Excel", bootstyle="primary", command=self.browse_files)
        self.choose_button.pack(pady=8)

        # Combobox chọn sheet
        self.sheet_dropdown = ttk.Combobox(main_frame, textvariable=self.selected_sheet, state='readonly', width=60, bootstyle="info")
        self.sheet_dropdown.pack(pady=10)

        # Nút dịch và lưu
        self.translate_button = ttk.Button(main_frame, text="🌍 Dịch và Lưu File", bootstyle="success", command=self.start_translate_thread)
        self.translate_button.pack(pady=12)

    def start_translate_thread(self):
        threading.Thread(target=self.translate_and_save, daemon=True).start()

    def browse_files(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx")])
        valid_files = [f for f in files if f.lower().endswith(".xlsx")]

        if not valid_files:
            Messagebox.show_error("Không có file hợp lệ.", title="Lỗi")
            return

        self.file_paths = valid_files[:10]
        self.files_label.config(text="\n".join([os.path.basename(f) for f in self.file_paths]))

        try:
            wb = load_workbook(self.file_paths[0])
            self.sheet_dropdown["values"] = ["Tất cả"] + wb.sheetnames
            self.selected_sheet.set("Tất cả")
        except:
            pass

    def start_translation_thread(self):
        if not self.file_paths:
            Messagebox.show_error("Vui lòng chọn ít nhất một file Excel.", title="Thiếu thông tin")
            return

        threading.Thread(target=self.translate_and_save, daemon=True).start()

    def translate_and_save(self):
        if not self.file_paths:
            Messagebox.show_error("Vui lòng chọn ít nhất một file Excel.", title="Thiếu thông tin")
            return

        # Giao diện popup loading
        progress_popup = tk.Toplevel(self.root)
        progress_popup.title("🔄 Đang xử lý")
        progress_popup.geometry("400x120")
        ttk.Label(progress_popup, text="🌐 Đang dịch các file, vui lòng chờ...", font=("Segoe UI", 11)).pack(pady=10)
        progress = ttk.Progressbar(progress_popup, length=300, mode='determinate', bootstyle="info-striped")
        progress.pack(pady=10)
        progress["maximum"] = len(self.file_paths)

        def task():
            for i, path in enumerate(self.file_paths):
                try:
                    wb = load_workbook(filename=path)
                    sheets_to_translate = wb.sheetnames if self.selected_sheet.get() == "Tất cả" else [self.selected_sheet.get()]

                    for sheet_name in sheets_to_translate:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows():
                            for cell in row:
                                if isinstance(cell.value, str):
                                    lines = re.split(r'\r?\n', cell.value)
                                    translated_lines = []
                                    for line in lines:
                                        stripped = line.strip()
                                        if stripped:
                                            translated = translate_dict.get(stripped, stripped)
                                            translated_lines.append(translated)
                                    cell.value = '\n'.join(translated_lines).strip()

                    base_dir = os.path.dirname(path)
                    base_name = os.path.splitext(os.path.basename(path))[0]
                    save_path = os.path.join(base_dir, f"{base_name}_translated.xlsx")
                    wb.save(save_path)

                except Exception as e:
                    print(f"Lỗi xử lý file {path}: {e}")

                # Cập nhật thanh tiến trình trong main thread
                self.root.after(0, lambda val=i+1: progress.configure(value=val))

            # Xong => dọn UI + báo thành công (tất cả trong main thread)
            def done_ui(self):
                self.file_paths.clear()
                self.sheet_dropdown['values'] = []
                self.selected_sheet.set("")
                self.drop_area.config(bootstyle="info")

                for widget in self.drop_area.winfo_children():
                    widget.destroy()

                self.drop_text = ttk.Label(self.drop_area, text="(Thả tối đa 10 file .xlsx vào đây)", font=("Segoe UI", 11))
                self.drop_text.pack(expand=True)                

                self.root.after(0, done_ui)

        # Chạy task trong luồng nền
        threading.Thread(target=task).start()

if __name__ == "__main__":
    app = Style().master
    ExcelTranslateApp(app)
    app.mainloop()