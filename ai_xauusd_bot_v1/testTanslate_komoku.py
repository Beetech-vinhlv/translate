import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinterdnd2 import DND_FILES, TkinterDnD
import os
from translator import Translator
from translation_dict import translate_dict

translator = Translator(translate_dict)

def translate_excel(file_path, sheet_names):
    try:
        wb = openpyxl.load_workbook(file_path)
        untranslated_values = set()

        def translate_cell_value(val):
            if not isinstance(val, str):
                return val
            cleaned = val.strip().replace('\n', '').replace('\r', '').replace('　', '').replace(' ', '')
            if cleaned in translate_dict:
                return translate_dict[cleaned]
            else:
                untranslated_values.add(cleaned)
                return val

        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = translate_cell_value(cell.value)

        if untranslated_values:
            print("\n🛠️ Các mục chưa có trong translate_dict:\n")
            for val in sorted(untranslated_values):
                print(f'"{val}": "{val}",')

        sheet_tag = "all" if len(sheet_names) > 1 else sheet_names[0]
        new_path = os.path.splitext(file_path)[0] + f"_translated_{sheet_tag}.xlsx"
        wb.save(new_path)
        return new_path

    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể xử lý file: {str(e)}")
        return None

def unified_translate_popup(file_path=None):
    selected_file = tk.StringVar(value=file_path or "")
    selected_sheets = []

    def browse_and_load():
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            selected_file.set(path)
            load_sheet_checkboxes(path)

    def load_sheet_checkboxes(path):
        try:
            wb = openpyxl.load_workbook(path)
            sheets = wb.sheetnames

            for widget in sheet_check_frame.winfo_children():
                widget.destroy()

            sheet_vars.clear()

            for sheet in sheets:
                var = tk.BooleanVar()
                chk = ttk.Checkbutton(sheet_check_frame, text=sheet, variable=var)
                chk.pack(anchor="w", padx=10)
                sheet_vars[sheet] = var

            all_checkbox_var.set(False)

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file: {str(e)}")

    def toggle_all_sheets():
        all_selected = all_checkbox_var.get()
        for var in sheet_vars.values():
            var.set(all_selected)

    def start_translate():
        path = selected_file.get()
        if not path:
            messagebox.showwarning("Thiếu file", "Vui lòng chọn file Excel.")
            return

        selected = [sheet for sheet, var in sheet_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning("Thiếu Sheet", "Vui lòng chọn ít nhất một sheet để dịch.")
            return

        result = translate_excel(path, selected)
        if result:
            messagebox.showinfo("✅ Thành công", f"Đã lưu file: {result}")
        popup.destroy()

    popup = tk.Toplevel()
    popup.title("📘 Dịch Excel")
    popup.geometry("420x600")
    popup.resizable(True, True)
    popup.grab_set()

    style = ttk.Style()
    style.configure("TLabel", font=("Segoe UI", 10))
    style.configure("TCheckbutton", font=("Segoe UI", 10))
    style.configure("TButton", font=("Segoe UI", 10, "bold"))

    container = ttk.Frame(popup, padding=20)
    container.pack(expand=True, fill=tk.BOTH)

    # File selector
    ttk.Label(container, text="📂 Chọn file Excel:").pack(anchor="w")
    file_frame = ttk.Frame(container)
    file_frame.pack(fill="x", pady=5)
    file_entry = ttk.Entry(file_frame, textvariable=selected_file, state="readonly")
    file_entry.pack(side="left", fill="x", expand=True)
    ttk.Button(file_frame, text="Browse", command=browse_and_load).pack(side="right", padx=5)

    # Sheet selector
    ttk.Label(container, text="📑 Chọn Sheet cần dịch:").pack(anchor="w", pady=(15, 5))
    all_checkbox_var = tk.BooleanVar()
    ttk.Checkbutton(container, text="Chọn tất cả", variable=all_checkbox_var, command=toggle_all_sheets).pack(anchor="w")

    sheet_check_frame = ttk.Frame(container)
    sheet_check_frame.pack(fill="both", expand=True, padx=10, pady=5)

    sheet_vars = {}

    if file_path:
        load_sheet_checkboxes(file_path)

    # Translate button
    ttk.Button(container, text="✅ Dịch và lưu", command=start_translate).pack(pady=20, fill="x")

    popup.wait_window()

def on_drop(event):
    file_path = event.data.strip('{}')
    if file_path.lower().endswith(".xlsx"):
        unified_translate_popup(file_path)
    else:
        messagebox.showwarning("Cảnh báo", "Vui lòng thả file .xlsx")

# Main GUI
app = TkinterDnD.Tk()
app.title("Excel Translator")
app.geometry("500x300")
app.resizable(False, False)

frame = ttk.Frame(app, padding=20)
frame.pack(expand=True, fill=tk.BOTH)

label = ttk.Label(frame, text="Kéo file Excel vào hoặc chọn file", font=("Arial", 12))
label.pack(pady=20)

drop_area = tk.Label(frame, text="🡇 KÉO FILE VÀO ĐÂY 🡇", relief="solid", width=40, height=8, background="white")
drop_area.pack(pady=10)
drop_area.drop_target_register(DND_FILES)
drop_area.dnd_bind('<<Drop>>', on_drop)

browse_btn = ttk.Button(frame, text="Chọn file", command=unified_translate_popup)
browse_btn.pack(pady=10)

app.mainloop()
