from flask import Flask, request, render_template, send_file, redirect
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from copy import copy
import os
import tempfile

app = Flask(__name__)

# Dummy function: thay bằng API hoặc translate_text thật
def translate_text(text):
    return text + " (vi)"

def process_excel(file_path):
    wb = load_workbook(file_path)
    out_wb = load_workbook(file_path)  # copy theo cách nhanh, giữ style
    for sheet in out_wb.worksheets:
        # Translate header (row 13, columns B-G)
        for col in range(2, 8):
            cell = sheet.cell(row=13, column=col)
            if isinstance(cell.value, str):
                cell.value = translate_text(cell.value.strip())

        # Translate "論理名" column below header
        header_row = None
        logical_col = None
        for r in range(1, 20):
            if sheet.cell(row=r, column=1).value == "No." and sheet.cell(row=r, column=2).value in ["論理名", "論理エンティティ名"]:
                header_row = r
                break
        if header_row:
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=header_row, column=col).value in ["論理名", "論理エンティティ名"]:
                    logical_col = col
                    break
            if logical_col:
                for r in range(header_row + 1, sheet.max_row + 1):
                    cell = sheet.cell(row=r, column=logical_col)
                    if isinstance(cell.value, str):
                        cell.value = translate_text(cell.value.strip())

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    out_wb.save(tmp.name)
    return tmp.name

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        uploaded_file = request.files["file"]
        if uploaded_file.filename.endswith((".xlsx", ".xls")):
            input_path = os.path.join(tempfile.gettempdir(), uploaded_file.filename)
            uploaded_file.save(input_path)
            output_path = process_excel(input_path)
            return send_file(output_path, as_attachment=True, download_name="translated.xlsx")
        else:
            return "Chỉ hỗ trợ file Excel (.xlsx, .xls)", 400
    return render_template("index.html")
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)

