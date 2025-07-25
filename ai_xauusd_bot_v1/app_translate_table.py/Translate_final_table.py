import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from copy import copy
from openpyxl.cell.cell import Cell

# Hàm dịch cột 論理名 và tên sheet sang tiếng Việt
def translate_text(japanese):
    translations = {
        # Dịch tên sheet
        "サマリ": "サマリ (Tóm tắt)",
        "バナーID": "バナーID (id panner)",
        "アイコンID": "アイコンID (id content)",
        "アクセス数": "アクセス数 (số access)",
        "登録者名": "登録者名 (tên người đăng ký)",
        "更新者名": "更新者名 (tên người update)",
        "QR画像ID": "QR画像ID (ID ảnh QR)",
        "画像URL": "画像URL (URL ảnh)",
        "表示開始日時": "表示開始日時 (ngày giờ bắt đầu hiển thị)",
        "表示終了日時": "表示終了日時 (ngày giờ kết thúc hiển thị)",
        "サムネイルID": "サムネイルID (ID thumbnail)",
        "エンティティ一覧": "エンティティ一覧 (Danh sách Entity)",
        "アクセスログテーブル": "アクセスログテーブル (Bảng nhật ký access)",
        "FMDNPユーザーマスタ": "FMDNPユーザーマスタ (Bảng chính người dùng FMDNP)",
        "利用規約マスタ": "利用規約マスタ (Bảng chính điều khoản sử dụng)",
        "口座申請": "口座申請 (request tài khoản)",
        "金融機関コードマスタ": "金融機関コードマスタ (Bảng chính mã tổ chức tài chính)",
        "お問い合わせマスタ": "お問い合わせマスタ (Bảng chính liên hệ)",
        "コンテンツマスタ": "コンテンツマスタ (Bảng chính nội dung)",
        "コンテンツNG理由": "コンテンツNG理由 (Lý do nội dung NG)",
        "コンテンツ申請": "コンテンツ申請 (request nội dung)",
        "コンテンツ申請印刷設定": "コンテンツ申請印刷設定 (Cài đặt in request nội dung)",
        "コンテンツ停止理由": "コンテンツ停止理由 (Lý do dừng nội dung)",
        "コンテンツ停止理由コメント": "コンテンツ停止理由コメント (Bình luận lý do dừng nội dung)",
        "コンテンツ停止解除理由": "コンテンツ停止解除理由 (Lý do hủy dừng nội dung)",
        "コンテンツ停止解除理由コメント": "コンテンツ停止解除理由コメント (Bình luận lý do hủy dừng nội dung)",
        "クリエイターマスタ": "クリエイターマスタ (Bảng chính creator)",
        "クリエイター活動NG理由": "クリエイター活動NG理由 (Lý do NG hoạt động creator)",
        "クリエイター活動申請": "クリエイター活動申請 (request hoạt động creator)",
        "クリエーター口座マスタ": "クリエーター口座マスタ (Bảng chính tài khoản creator)",
        "携帯電話番号再設定": "携帯電話番号再設定 (Đặt lại số điện thoại di động)",
        "クリエイターマイページお知らせマスタ": "クリエイターマイページお知らせマスタ (Bảng chính thông báo trang cá nhân creator)",
        "初回クリエイター名申請": "初回クリエイター名申請 (request tên creator lần đầu)",
        "クリエイター通知": "クリエイター通知 (Thông báo creator)",
        "クリエイターワンタイム携帯電話番号認証": "クリエイターワンタイム携帯電話番号認証 (Xác thực số điện thoại di động một lần creator)",
        "クリエーターパスワード再設定": "クリエーターパスワード再設定 (Đặt lại mật khẩu creator)",
        "クリエーター本人情報マスタ": "クリエーター本人情報マスタ (Bảng chính thông tin cá nhân creator)",
        "クリエイター本人情報審査": "クリエイター本人情報審査 (Xét duyệt thông tin cá nhân creator)",
        "クリエイター本人情報審査NG理由": "クリエイター本人情報審査NG理由 (Lý do NG xét duyệt thông tin cá nhân creator)",
        "クリエイタープロフィールマスタ": "クリエイタープロフィールマスタ (Bảng chính hồ sơ creator)",
        "クリエイター申請": "クリエイター申請 (request creator)",
        "クリエイター利用規約許諾": "クリエイター利用規約許諾 (Đồng ý điều khoản sử dụng creator)",
        "月次クリエイター毎支払い管理": "月次クリエイター毎支払い管理 (Quản lý thanh toán hàng tháng theo creator)",
        "公式サイトお知らせマスタ": "公式サイトお知らせマスタ (Bảng chính thông báo trang chính thức)",
        "NGチェック項目マスタ": "NGチェック項目マスタ (Bảng chính hạng mục kiểm tra NG)",
        "支払い": "支払い (Thanh toán)",
        "清算明細": "清算明細 (Chi tiết thanh toán)",
        "支払い管理": "支払い管理 (Quản lý thanh toán)",
        "郵便番号マスタ": "郵便番号マスタ (Bảng chính mã bưu điện)",
        "印刷実績": "印刷実績 (Kết quả in)",
        "印刷設定マスタ": "印刷設定マスタ (Bảng chính cài đặt in)",
        "プリント番号印刷設定": "プリント番号印刷設定 (Cài đặt in mã print)",
        "差戻停止Ful歴": "差戻停止履歴 (Lịch sử từ chối và dừng)",
        "販売申請": "販売申請 (request bán hàng)",
        "電話番号申請": "電話番号申請 (request số điện thoại)",
        "解約申請": "解約申請 (request hủy hợp đồng)",
        "T番号マスタ": "T番号マスタ (Bảng chính số T)",
        "T番号審査": "T番号審査 (Xét duyệt số T)",
        "T番号審査NG理由": "T番号審査NG理由 (Lý do NG xét duyệt số T)",
        # Dịch cột 論理名 và 論理エンティティ名
        "ID": "ID",
        "論理名": "tên logic",
        "物理名": "tên vật lý",
        "データ型": "kiểu dữ liệu",
        "デフォルト": "default",
        "アクセス日時": "アクセス日時 (Ngày giờ access)",
        "ユーザー区分": "ユーザー区分 (Phân loại người dùng)",
        "ユーザーID": "ユーザーID (user_id)",
        "ユーザー名": "ユーザー名 (Tên người dùng)",
        "画面名": "画面名 (Tên màn hình)",
        "アクセス内容": "アクセス内容 (Nội dung access)",
        "備考": "備考 (note)",
        "登録者ID": "登録者ID (ID người tạo)",
        "登録日時": "登録日時 (Ngày giờ tạo)",
        "更新者ID": "更新者ID (ID người cập nhật)",
        "更新日時": "更新日時 (Ngày giờ cập nhật)",
        "メールアドレス": "メールアドレス (Địa chỉ email)",
        "パスワード": "パスワード (Mật khẩu)",
        "ロール": "ロール (role)",
        "パスワード有効期限": "パスワード有効期限 (Thời hạn mật khẩu)",
        "ログイントークン": "ログイントークン (login token)",
        "ステータス": "ステータス (status)",
        "削除日時": "削除日時 (Ngày giờ xóa)",
        "本文": "本文 (body)",
        "変更確認必須規約": "変更確認必須規約 (Quy định bắt buộc xác nhận thay đổi)",
        "口座審査番号": "口座審査番号 (Số xét duyệt tài khoản)",
        "クリエイター申請ID": "クリエイター申請ID (ID request creator)",
        "金融機関": "金融機関 (Tổ chức tài chính)",
        "支店コード": "支店コード (Mã chi nhánh)",
        "口座の種類": "口座の種類 (Loại tài khoản)",
        "ゆうちょ番号フラグ": "ゆうちょ番号フラグ (flag số bưu điện)",
        "ゆうちょ番号": "ゆうちょ番号 (Số bưu điện)",
        "口座番号": "口座番号 (Số tài khoản)",
        "口座名義": "口座名義 (Tên chủ tài khoản)",
        "NGチェックID": "NGチェックID (ID kiểm tra NG)",
        "業態": "業態 (Loại hình kinh doanh)",
        "金融機関等名称": "金融機関等名称 (Tên tổ chức tài chính)",
        "店舗名称": "店舗名称 (Tên cửa hàng)",
        "金融機関等コード": "金融機関等コード (Mã tổ chức tài chính)",
        "店舗コード": "店舗コード (Mã cửa hàng)",
        "日銀当座勘定取引店": "日銀当座勘定取引店 (Cửa hàng giao dịch tài khoản hiện tại BOJ)",
        "Toユーザー種別": "Toユーザー種別 (Loại người dùng nhận)",
        "Fromユーザー種別": "Fromユーザー種別 (Loại người dùng gửi)",
        "返信用メールアドレス": "返信用メールアドレス (Địa chỉ email trả lời)",
        "件名": "件名 (Tiêu đề)",
        "内容": "内容 (Nội dung)",
        "複合機用ID": "複合機用ID (ID dùng cho máy in đa năng)",
        "クリエイターID": "クリエイターID (ID creator)",
        "商品名": "商品名 (Tên sản phẩm)",
        "複合機印刷用画像URL": "複合機印刷用画像URL (URL hình ảnh in đa năng)",
        "公式サイトサムネイルURL": "公式サイトサムネイルURL (URL hình thu nhỏ trang chính thức)",
        "QR画像URL": "QR画像URL (URL hình ảnh QR)",
        "公開開始日時": "公開開始日時 (Ngày giờ bắt đầu công khai)",
        "公開終了日時": "公開終了日時 (Ngày giờ kết thúc công khai)",
        "NG対象": "NG対象 (Đối tượng NG)",
        "コンテンツ申請ID": "コンテンツ申請ID (ID request nội dung)",
        "コンテンツID": "コンテンツID (ID nội dung)",
        "販売申請ID": "販売申請ID (ID request bán hàng)",
        "印刷設定ID": "印刷設定ID (ID cài đặt in)",
        "停止理由コメント": "停止理由コメント (Bình luận lý do dừng)",
        "停止解除理由コメント": "停止解除理由コメント (Bình luận lý do hủy dừng)",
        "クリエイター識別子": "クリエイター識別子 (Định danh creator)",
        "変更利用規約同意": "変更利用規約同意 (Đồng ý thay đổi điều khoản sử dụng)",
        "最終ログイン日時": "最終ログイン日時 (Ngày giờ đăng nhập cuối cùng)",
        "ログイン試行回数": "ログイン試行回数 (Số lần thử đăng nhập)",
        "ログイン解除日時": "ログイン解除日時 (Ngày giờ mở khóa đăng nhập)",
        "更新試行回数": "更新試行回数 (Số lần thử cập nhật)",
        "更新ロック解除日時": "更新ロック解除日時 (Ngày giờ mở khóa cập nhật)",
        "クリエイター活動申請ID": "クリエイター活動申請ID (ID request hoạt động creator)",
        "クリエイター名": "クリエイター名 (Tên creator)",
        "バナーURL": "バナーURL (URL biểu ngữ)",
        "アイコンURL": "アイコンURL (URL biểu tượng)",
        "携帯電話番号": "携帯電話番号 (Số điện thoại di động)",
        "ワンタイムコード": "ワンタイムコード (Mã một lần)",
        "ワンタイムコード有効期限": "ワンタイムコード有効期限 (Thời hạn mã một lần)",
        "試行回数": "試行回数 (Số lần thử)",
        "トークン": "トークン (Mã thông báo)",
        "トークン有効期限": "トークン有効期限 (Thời hạn mã thông báo)",
        "認証フラグ": "認証フラグ (flag xác thực)",
        "ロック解除日時": "ロック解除日時 (Ngày giờ mở khóa)",
        "生成日時": "生成日時 (Ngày giờ tạo)",
        "氏名": "氏名 (Họ tên)",
        "カナ": "カナ (Kana)",
        "生年月日": "生年月日 (Ngày sinh)",
        "郵便番号": "郵便番号 (Mã bưu điện)",
        "都道府県名": "都道府県名 (Tên tỉnh)",
        "市区町村": "市区町村 (Thành phố/Quận/Huyện)",
        "丁・番地": "丁・番地 (Địa chỉ lô/số nhà)",
        "建物名・部屋番号": "建物名・部屋番号 (Tên tòa nhà/số phòng)",
        "審査番号": "審査番号 (Số xét duyệt)",
        "ログインURL": "ログインURL (URL đăng nhập)",
        "クリエイター本人情報審査ID": "クリエイター本人情報審査ID (ID xét duyệt thông tin cá nhân creator)",
        "電話番号": "電話番号 (Số điện thoại)",
        "アクセス管理者ID": "アクセス管理者ID (ID quản trị viên access)",
        "バージョン": "バージョン (version)",
        "利用規約ID": "利用規約ID (ID điều khoản sử dụng)",
        "該当月": "該当月 (Tháng tương ứng)",
        "支払い予定日": "支払い予定日 (Ngày thanh toán dự kiến)",
        "支払い件数": "支払い件数 (Số lượng thanh toán)",
        "重要通知フラグ": "重要通知フラグ (flag thông báo quan trọng)",
        "重要通知表示フラグ": "重要通知表示フラグ (flag hiển thị thông báo quan trọng)",
        "表示フラグ": "表示フラグ (flag hiển thị)",
        "ソート順": "ソート順 (Thứ tự sort)",
        "表示文言": "表示文言 (text hiển thị)",
        "通知件名": "通知件名 (title thông báo)",
        "通知本文": "通知本文 (body thông báo)",
        "請求管理番号": "請求管理番号 (Số quản lý hóa đơn)",
        "支払い金額(税込)": "支払い金額(税込) (Số tiền thanh toán bao gồm thuế)",
        "支払い金額(税抜)": "支払い金額(税抜) (Số tiền thanh toán không bao gồm thuế)",
        "消費税等": "消費税等 (Thuế tiêu thụ)",
        "源泉徴収税率": "源泉徴収税率 (Tỷ lệ thuế khấu trừ tại nguồn)",
        "源泉徴収税": "源泉徴収税 (Thuế khấu trừ tại nguồn)",
        "対象開始日": "対象開始日 (Ngày bắt đầu áp dụng)",
        "対象終了日": "対象終了日 (Ngày kết thúc áp dụng)",
        "支払日": "支払日 (Ngày thanh toán)",
        "銀行名": "銀行名 (Tên ngân hàng)",
        "支店名": "支店名 (Tên chi nhánh)",
        "印刷設定": "印刷設定 (Cài đặt in)",
        "プリント番号": "プリント番号 (mã print)",
        "単価": "単価 (Đơn giá)",
        "差し戻し": "差し戻し (reject)",
        "数量": "数量 (Số lượng)",
        "小計": "小計 (Tổng phụ)",
        "全国地方公共団体コード": "全国地方公共団体コード (Mã cơ quan chính quyền địa phương toàn quốc)",
        "(旧)郵便番号(5桁)": "(旧)郵便番号(5桁) (Mã bưu điện cũ 5 chữ số)",
        "郵便番号(7桁)": "郵便番号(7桁) (Mã bưu điện 7 chữ số)",
        "都道府件名(カタカナ)": "都道府件名(カタカナ) (Tên tỉnh bằng Kana)",
        "市区町村(カタカナ)": "市区町村(カタカナ) (Tên thành phố/quận/huyện bằng Kana)",
        "町域名(カタカナ)": "町域名(カタカナ) (Tên khu vực bằng Kana)",
        "町域名": "町域名 (Tên khu vực)",
        "一町域が二以上の郵便番号で表されるフラグ": "一町域が二以上の郵便番号で表されるフラグ (flag khu vực được biểu thị bằng hai mã bưu điện trở lên)",
        "小字毎に番地が起番されている町域フラグ": "小字毎に番地が起番されている町域フラグ (flag khu vực có số địa chỉ được đánh số theo từng đơn vị nhỏ)",
        "丁目を有する町域フラグ": "丁目を有する町域フラグ (flag khu vực có lô)",
        "一つの郵便番号で二以上の町域を表すフラグ": "一つの郵便番号で二以上の町域を表すフラグ (flag một mã bưu điện biểu thị hai khu vực trở lên)",
        "更新区分": "更新区分 (Phân loại cập nhật)",
        "変更理由区分": "変更理由区分 (Phân loại lý do thay đổi)",
        "利用日時": "利用日時 (Ngày giờ sử dụng)",
        "コンテンツ名": "コンテンツ名 (Tên nội dung)",
        "カラーモード": "カラーモード (Chế độ màu)",
        "用紙サイズ": "用紙サイズ (Kích thước giấy)",
        "単価(税込)": "単価(税込) (Đơn giá bao gồm thuế)",
        "使用回数": "使用回数 (Số lần sử dụng)",
        "価格": "価格 (Giá)",
        "両面設定": "両面設定 (Cài đặt in hai mặt)",
        "販売情報審査番号": "販売情報審査番号 (Số xét duyệt thông tin bán hàng)",
        "申請日": "申請日 (Ngày request)",
        "アクセスユーザー": "アクセスユーザー (Người dùng access)",
        "解約申請番号": "解約申請番号 (Số request hủy hợp đồng)",
        "T番号": "T番号 (Số T)",
        "事業者名": "事業者名 (Tên doanh nghiệp)",
        "事業登記住所": "事業登記住所 (Địa chỉ đăng ký kinh doanh)",
        "T番号審査ID": "T番号審査ID (ID xét duyệt số T)",
        "通知区分": "通知区分 (Phân loại thông báo)",
        "NG通知対象": "NG通知対象 (Đối tượng thông báo NG)",
        "申請ID": "申請ID (ID request)",
        "論理エンティティ名": "論理エンティティ名 (Tên logic Entity)",
        "物理エンティティ名": "物理エンティティ名 (Tên vật lý Entity)",
        "タグ": "タグ (thẻ)"
    }
    if japanese in translations:
        return translations[japanese]
    else:
        # Có thể log tại đây nếu cần ghi lại những từ chưa dịch
        # print(f"⚠️ Không có bản dịch cho: {japanese}")
        return japanese  # Giữ nguyên nếu không có bản dịch

# Định dạng hàng tiêu đề và cột
def format_sheet(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "No." and row[1].value == "論理名":
            for cell in row:
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            break

    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        if col[0].value in ["No.", "Not Null", "デフォルト"]:
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            for cell in col:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    column_widths = {'A': 5, 'B': 20, 'C': 15, 'D': 30, 'E': 10, 'F': 10, 'G': 50}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

# Hàm xử lý dịch
def translate_excel(input_path, on_complete_callback, progress_window):
    try:
        input_wb = openpyxl.load_workbook(input_path)
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể mở file: {e}")
        progress_window.destroy()
        return

    output_wb = openpyxl.Workbook()
    output_wb.remove(output_wb.active)

    total_sheets = len(input_wb.sheetnames)

    for idx, sheet_name in enumerate(input_wb.sheetnames, 1):
        progress_window.nametowidget("progress_label").config(
            text=f"Đang dịch sheet: {sheet_name} ({idx}/{total_sheets})"
        )
        progress_window.nametowidget("progress_bar")['value'] = (idx / total_sheets) * 100
        progress_window.update()

        input_sheet = input_wb[sheet_name]
        # 👉 KHÔNG dịch tên sheet nữa, giữ nguyên
        output_sheet = output_wb.create_sheet(title=sheet_name)

        # Copy toàn bộ nội dung và định dạng
        for row in input_sheet.iter_rows():
            for cell in row:
                if not isinstance(cell, Cell):
                    continue
                new_cell = output_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copy chiều rộng cột
        for col_letter, col_dim in input_sheet.column_dimensions.items():
            output_sheet.column_dimensions[col_letter].width = col_dim.width

        # Copy merged cells
        for merged_cell_range in input_sheet.merged_cells.ranges:
            output_sheet.merge_cells(str(merged_cell_range))

        # 👉 DỊCH các ô dòng 13 từ cột B~G
        for col in range(2, 8):  # B=2 đến G=7
            cell = output_sheet.cell(row=13, column=col)
            if isinstance(cell.value, str) and cell.value.strip():
                cell.value = translate_text(cell.value.strip())

        # 👉 Tìm dòng tiêu đề & cột cần dịch dữ liệu
        header_row = None
        logical_name_col = None

        for row in range(1, 20):
            col1_val = input_sheet.cell(row=row, column=1).value
            col2_val = input_sheet.cell(row=row, column=2).value
            if col1_val == "No." and col2_val in ["論理名", "論理エンティティ名"]:
                header_row = row
                break
        if header_row is None:
            continue

        for col in range(1, input_sheet.max_column + 1):
            val = input_sheet.cell(row=header_row, column=col).value
            if val in ["論理名", "論理エンティティ名"]:
                logical_name_col = col
                break
        if logical_name_col is None:
            continue

        # 👉 DỊCH dữ liệu trong cột "論理名"
        for row in range(header_row + 1, input_sheet.max_row + 1):
            cell = output_sheet.cell(row=row, column=logical_name_col)
            original_value = cell.value
            if isinstance(original_value, str):
                cell.value = translate_text(original_value.strip())

    progress_window.destroy()

    output_path = filedialog.asksaveasfilename(
        title="Chọn nơi lưu file Excel đã dịch",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if output_path:
        output_wb.save(output_path)
        messagebox.showinfo("Hoàn tất", f"✅ Dịch thành công!\nFile đã lưu: {output_path}")
        on_complete_callback()
    else:
        messagebox.showwarning("Hủy", "Bạn chưa chọn nơi lưu file. Dữ liệu chưa được lưu.")

# Giao diện chính
def run_gui():
    def on_select_file():
        input_path = filedialog.askopenfilename(
            title="Chọn file Excel gốc",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not input_path:
            return

        # Hiện modal loading
        progress_window = tk.Toplevel(root)
        progress_window.title("Đang xử lý...")
        progress_window.geometry("400x100")
        progress_window.resizable(False, False)
        tk.Label(progress_window, text="Đang bắt đầu dịch...").pack(pady=10)

        progress_label = tk.Label(progress_window, name="progress_label", text="")
        progress_label.pack()

        progress_bar = ttk.Progressbar(progress_window, name="progress_bar", mode="determinate", length=350)
        progress_bar.pack(pady=5)

        # Bắt đầu dịch ở thread riêng để không bị treo giao diện
        threading.Thread(
            target=lambda: translate_excel(input_path, on_complete_callback=lambda: None, progress_window=progress_window),
            daemon=True
        ).start()

    root = tk.Tk()
    root.title("Dịch Excel Nhật-Việt")
    root.geometry("400x180")
    root.resizable(False, False)

    tk.Label(root, text="Chọn file Excel cần dịch", font=("Arial", 12)).pack(pady=20)
    tk.Button(root, text="📁 Chọn file", command=on_select_file, font=("Arial", 11), width=20).pack(pady=10)

    tk.Label(root, text="Sau khi dịch xong, sẽ chọn nơi lưu", fg="gray").pack(pady=5)

    root.mainloop()

# Chạy chương trình
if __name__ == "__main__":
    run_gui()