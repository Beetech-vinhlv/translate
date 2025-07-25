import os
import re
import tkinter as tk
import traceback
from tkinter import filedialog, messagebox
from tkinterdnd2 import TkinterDnD, DND_FILES
from openpyxl import load_workbook
from ttkbootstrap import Style
from ttkbootstrap import ttk
from ttkbootstrap.dialogs import Messagebox
from translator import Translator
from translation_dict import translate_dict

translator = Translator(translate_dict)
class ExcelTranslateApp:
    def __init__(self, root):
        self.root = root
        self.root.title("✨ Excel Translator ✨")
        self.root.geometry("750x350")
        self.root.resizable(True, True)

        self.file_paths = []  # Danh sách nhiều file
        self.selected_sheet = tk.StringVar()

        self.setup_widgets()

    def setup_widgets(self):
        frame = ttk.Frame(self.root, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="📁 Kéo thả hoặc chọn tối đa 10 file Excel", font=("Segoe UI", 12)).pack(pady=8)

        self.drop_frame = tk.Frame(frame, height=100, width=700, bg="#e0f7fa", relief=tk.RIDGE, bd=2)
        self.drop_frame.pack(pady=5)
        self.drop_frame.pack_propagate(False)

        drop_label = ttk.Label(self.drop_frame, text="(Kéo thả file .xlsx vào đây)", font=("Segoe UI", 11))
        drop_label.pack(expand=True)

        self.drop_frame.drop_target_register(DND_FILES)
        self.drop_frame.dnd_bind('<<Drop>>', self.on_drop_files)

        ttk.Button(frame, text="📂 Chọn file", bootstyle="primary", command=self.browse_files).pack(pady=5)

        self.sheet_dropdown = ttk.Combobox(frame, textvariable=self.selected_sheet, state='readonly', width=60, bootstyle="info")
        self.sheet_dropdown.pack(pady=10)

        ttk.Button(frame, text="🌍 Dịch và Lưu File", bootstyle="success", command=self.translate_and_save).pack(pady=10)

    def browse_files(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx")])
        self.add_files(files)

    def on_drop_files(self, event):
        raw_files = self.root.tk.splitlist(event.data)
        self.add_files(raw_files)

    def add_files(self, files):
        valid_files = [f for f in files if f.lower().endswith(".xlsx")]
        if not valid_files:
            messagebox.showerror("Lỗi", "Không tìm thấy file .xlsx hợp lệ.")
            return

        self.file_paths.extend(valid_files)
        self.file_paths = list(set(self.file_paths))[:10]  # Giới hạn 10 file, không trùng

        files_display = "\n".join([os.path.basename(f) for f in self.file_paths])
        self.drop_frame.config(bg="#c8e6c9")
        for widget in self.drop_frame.winfo_children():
            widget.destroy()
        ttk.Label(self.drop_frame, text=f"✅ {len(self.file_paths)} file đã chọn:\n{files_display}",
                  font=("Segoe UI", 10), justify="center").pack(expand=True)

        if self.file_paths:
            try:
                wb = load_workbook(filename=self.file_paths[0])
                sheet_names = wb.sheetnames
                self.sheet_dropdown['values'] = ['Tất cả'] + sheet_names
                self.selected_sheet.set('Tất cả')
            except:
                pass

    def get_unique_filename(self, directory, base_name, extension):
        filename = f"{base_name}{extension}"
        counter = 1
        while os.path.exists(os.path.join(directory, filename)):
            filename = f"{base_name}_{counter}{extension}"
            counter += 1
        return os.path.join(directory, filename)
    
    def translate_and_save(self):
        if not self.file_paths:
            messagebox.showerror("Thiếu thông tin", "Vui lòng chọn ít nhất một file Excel.")
            return

        # Tạo popup hiển thị tiến độ
        progress_popup = tk.Toplevel(self.root)
        progress_popup.title("⏳ Đang dịch Excel...")
        progress_popup.geometry("480x160")
        progress_popup.configure(bg="#f7f9fc")
        progress_popup.resizable(False, False)

        ttk.Label(progress_popup, text="⏳ Đang dịch, vui lòng chờ...", font=("Segoe UI", 11, "bold")).pack(pady=(15, 5))
        progress_label = ttk.Label(progress_popup, text="", font=("Segoe UI", 10), background="#f7f9fc")
        progress_label.pack()

        progress_bar = ttk.Progressbar(
            progress_popup, length=380, mode='determinate', style="TProgressbar"
        )
        progress_bar.pack(pady=(10, 5))
        progress_bar["maximum"] = len(self.file_paths)

        self.root.update_idletasks()

        selected_sheet = self.selected_sheet.get()

        for i, path in enumerate(self.file_paths):
            try:
                wb = load_workbook(filename=path)
                sheets_to_translate = wb.sheetnames if selected_sheet == "Tất cả" else [selected_sheet]

                for idx, sheet_name in enumerate(sheets_to_translate, 1):
                    progress_label.config(
                        text=f"📄 {os.path.basename(path)}\nSheet: {sheet_name} ({idx}/{len(sheets_to_translate)})"
                    )
                    progress_popup.update()

                    ws = wb[sheet_name]
                    for row in ws.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str):
                                # Tách dòng, dịch và gộp lại nhanh hơn bằng list comprehension
                                translated_lines = [
                                    translate_dict.get(line.strip(), line.strip())
                                    for line in re.split(r'\r?\n', cell.value) if line.strip()
                                ]
                                cell.value = '\n'.join(translated_lines)

                # Lưu file với tên không trùng lặp
                base_dir = os.path.dirname(path)
                base_name = os.path.splitext(os.path.basename(path))[0] + "_translated"
                save_path = self.get_unique_filename(base_dir, base_name, ".xlsx")
                wb.save(save_path)

            except Exception:
                print(f"Lỗi dịch {path}:\n{traceback.format_exc()}")

            progress_bar["value"] = i + 1
            progress_popup.update()

        progress_popup.destroy()
        messagebox.showinfo("✅ Dịch hoàn tất", f"Đã xử lý {len(self.file_paths)} file thành công.")

        # Reset giao diện sau khi hoàn tất
        self.file_paths.clear()
        self.sheet_dropdown['values'] = []
        self.selected_sheet.set("")
        self.drop_frame.config(bg="#e0f7fa")

        for widget in self.drop_frame.winfo_children():
            widget.destroy()
        ttk.Label(self.drop_frame, text="(Kéo thả file .xlsx vào đây)", font=("Segoe UI", 11)).pack(expand=True)


if __name__ == "__main__":
    app = TkinterDnD.Tk()
    style = Style("flatly")
    ExcelTranslateApp(app)
    app.mainloop()